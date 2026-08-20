# -*- coding: utf-8 -*-
"""
Tạo/ghi đè sheet 'chi_so_tai_chinh' (sheet thứ 4) trong mỗi file
financials/<MA>/<MA>_financials.xlsx.

Toàn bộ số liệu trong sheet này được lấy bằng công thức INDEX/MATCH,
dò theo TÊN khoản mục (cột A, sheet balance_sheet/income_statement) và
theo NĂM (dòng 1 của các sheet nguồn) — không tham chiếu ô cố định.
Vì vậy sheet vẫn chạy đúng dù:
    - thứ tự dòng trong balance_sheet/income_statement thay đổi
    - số lượng năm (cột) tăng lên qua các lần fetch mới
Nhãn khoản mục ở cột A của khối "DỮ LIỆU TRÍCH XUẤT" được lấy ĐÚNG theo tên
gốc xuất hiện trong báo cáo tài chính của từng công ty (theo mapping ứng với
từng loại hình DN — STANDARD/BANK/SECURITIES) để người đọc đối chiếu ngay
được với dữ liệu gốc; công thức ở khối "CHỈ SỐ TÀI CHÍNH" tự dò lại theo
nhãn đó, không phụ thuộc thứ tự dòng.

Bố cục sheet (trên xuống): tiêu đề -> khối CHỈ SỐ TÀI CHÍNH (kết quả) ->
khối DỮ LIỆU TRÍCH XUẤT (nguồn, tham chiếu). Năm được xếp TĂNG DẦN từ trái
sang phải (năm cũ nhất bên trái, mới nhất bên phải), khớp với thứ tự cột
trong 3 sheet báo cáo nguồn (balance_sheet/income_statement/cash_flow) —
2 bên đã đồng bộ nên header năm ở sheet này lấy thẳng theo cột tương ứng,
không cần đảo. Phần INDEX/MATCH dò theo GIÁ TRỊ năm nên vẫn đúng dù thứ tự
cột trong sheet nguồn có thay đổi sau này.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financials")
SHEET_NAME = "chi_so_tai_chinh"

# Số cột năm (YEAR_COLS/REVERSE_COL/PRIOR_COL) KHÔNG còn khai báo cố định ở
# đây nữa — được tính động trong build_workbook_ratio_sheet() theo đúng số
# năm thực có ở dòng 1 sheet balance_sheet của từng công ty (để không bỏ sót
# năm nào khi có dữ liệu mới). Thứ tự vẫn xếp TĂNG DẦN trái->phải (cũ nhất
# bên trái, mới nhất bên phải), khớp thứ tự cột trong sheet nguồn.

FONT_NAME = "Tahoma"
GREEN = "008000"    # link sang sheet khác (khối trích xuất)
BLACK = "000000"    # công thức nội bộ (khối chỉ số)
GREY = "808080"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tô màu theo giá trị (tham khảo cách làm ở file FPT mẫu: fill nền theo
# ngưỡng giá trị bằng Conditional Formatting "Cell Is").
FILL_GOOD = PatternFill("solid", fgColor="C6E0B4")   # xanh nhạt: tích cực
FILL_BAD = PatternFill("solid", fgColor="F8CBAD")    # cam/đỏ nhạt: rủi ro/tiêu cực
FILL_AMBER = PatternFill("solid", fgColor="FFE699")  # vàng/hổ phách: trung tính/cảnh báo nhẹ

# ---------------------------------------------------------------------------
# 1) Khai báo khoản mục thô cần trích xuất: key -> nhãn hiển thị mặc định
#    (chỉ dùng khi công ty không có khoản mục tương ứng / hiển thị "n/a";
#    nếu có mapping, nhãn thực tế lấy đúng theo tên gốc trong báo cáo — xem
#    template bên dưới).
# ---------------------------------------------------------------------------
RAW_ITEMS = [
    ("equity",           "Vốn chủ sở hữu"),
    ("total_assets",     "Tổng tài sản"),
    ("total_liab",       "Nợ phải trả"),
    ("st_liab",          "Nợ ngắn hạn"),
    ("lt_liab",          "Nợ dài hạn"),
    ("st_borrow",        "Vay và nợ thuê tài chính ngắn hạn"),
    ("lt_borrow",        "Vay và nợ thuê tài chính dài hạn"),
    ("revenue",          "Doanh thu thuần"),
    ("cogs",             "Giá vốn hàng bán / Chi phí hoạt động"),
    ("gross_profit",     "Lợi nhuận gộp"),
    ("selling_exp",      "Chi phí bán hàng"),
    ("admin_exp",        "Chi phí quản lý"),
    ("operating_profit", "Lợi nhuận từ hoạt động kinh doanh"),
    ("interest_exp",     "Chi phí lãi vay"),
    ("ni_total",         "Lợi nhuận sau thuế (toàn công ty)"),
    ("ni_parent",        "Lợi nhuận của Cổ đông Công ty mẹ"),
    ("ar",               "Phải thu khách hàng"),
    ("ap",               "Phải trả người bán"),
    ("inventory",        "Hàng tồn kho, ròng"),
    ("cash",             "Tiền và tương đương tiền"),
]

# ---------------------------------------------------------------------------
# 2) Mapping (sheet nguồn, nhãn CHÍNH XÁC trong cột A của sheet đó) theo từng
#    kiểu báo cáo. None = khoản mục không tồn tại / không có ý nghĩa với kiểu
#    doanh nghiệp này -> sheet sẽ hiển thị "n/a".
# ---------------------------------------------------------------------------
BS, IS_ = "balance_sheet", "income_statement"

TEMPLATE_STANDARD = {
    "equity":           (BS, "Vốn chủ sở hữu"),
    "total_assets":      (BS, "TỔNG CỘNG TÀI SẢN"),
    "total_liab":        (BS, "NỢ PHẢI TRẢ"),
    "st_liab":           (BS, "Nợ ngắn hạn"),
    "lt_liab":           (BS, "Nợ dài hạn"),
    "st_borrow":         (BS, "Vay ngắn hạn"),
    "lt_borrow":         (BS, "Vay dài hạn"),
    "revenue":           (IS_, "Doanh thu thuần"),
    "cogs":              (IS_, "Giá vốn hàng bán"),
    "gross_profit":      (IS_, "Lợi nhuận gộp"),
    "selling_exp":       (IS_, "Chi phí bán hàng"),
    "admin_exp":         (IS_, "Chi phí quản lý doanh nghiệp"),
    "operating_profit":  (IS_, "Lãi/(lỗ) từ hoạt động kinh doanh"),
    "interest_exp":      (IS_, "Chi phí lãi vay"),
    "ni_total":          (IS_, "Lãi/(lỗ) thuần sau thuế"),
    "ni_parent":         (IS_, "Lợi nhuận của Cổ đông của Công ty mẹ"),
    "ar":                (BS, "Phải thu khách hàng"),
    "ap":                (BS, "Phải trả người bán"),
    "inventory":         (BS, "Hàng tồn kho, ròng"),
    "cash":               (BS, "Tiền và tương đương tiền"),
}

TEMPLATE_BANK = {
    "equity":            (BS, "VỐN CHỦ SỞ HỮU"),
    "total_assets":      (BS, "TỔNG TÀI SẢN"),
    "total_liab":        (BS, "TỔNG NỢ PHẢI TRẢ"),
    "st_liab":           None,
    "lt_liab":           None,
    "st_borrow":         None,
    "lt_borrow":         None,
    "revenue":           (IS_, "Tổng thu nhập hoạt động"),
    "cogs":              None,
    "gross_profit":      None,
    "selling_exp":       None,
    "admin_exp":         (IS_, "Chi phí quản lý doanh nghiệp"),
    "operating_profit":  (IS_, "Lợi nhuận thuần hoạt động trước khi trích lập dự phòng tổn thất tín dụng"),
    "interest_exp":      (IS_, "Chi phí lãi và các chi phí tương tự"),
    "ni_total":          (IS_, "Lợi nhuận sau thuế"),
    "ni_parent":         (IS_, "Cổ đông của Công ty mẹ"),
    "ar":                None,
    "ap":                None,
    "inventory":         None,
    "cash":               (BS, "Tiền mặt, vàng bạc, đá quý"),
}

TEMPLATE_SECURITIES = {
    "equity":            (BS, "Vốn chủ sở hữu"),
    "total_assets":      (BS, "TỔNG CỘNG TÀI SẢN"),
    "total_liab":        (BS, "NỢ PHẢI TRẢ"),
    "st_liab":           (BS, "Nợ phải trả ngắn hạn"),
    "lt_liab":           (BS, "Nợ phải trả dài hạn"),
    "st_borrow":         (BS, "Vay ngắn hạn"),
    "lt_borrow":         (BS, "Vay dài hạn"),
    "revenue":           (IS_, "Doanh thu thuần về hoạt động kinh doanh"),
    "cogs":              (IS_, "CHI PHÍ HOẠT ĐỘNG"),
    "gross_profit":      (IS_, "LỢI NHUẬN GỘP"),
    "selling_exp":       (IS_, "CHI PHÍ BÁN HÀNG"),
    "admin_exp":         (IS_, "CHI PHÍ QUẢN LÝ CÔNG TY CHỨNG KHOÁN"),
    "operating_profit":  (IS_, "KẾT QUẢ HOẠT ĐỘNG"),
    "interest_exp":      (IS_, "Chi phí lãi vay"),
    "ni_total":          (IS_, "LỢI NHUẬN KẾ TOÁN SAU THUẾ"),
    "ni_parent":         (IS_, "Lợi nhuận sau thuế phân bổ cho chủ sở hữu"),
    "ar":                None,
    "ap":                (BS, "Phải trả người bán ngắn hạn"),
    "inventory":         None,
    "cash":               (BS, "Tiền và tương đương tiền"),
}

SYMBOL_TEMPLATE = {
    "HPG": ("STANDARD", TEMPLATE_STANDARD),
    "MWG": ("STANDARD", TEMPLATE_STANDARD),
    "PNJ": ("STANDARD", TEMPLATE_STANDARD),
    "FRT": ("STANDARD", TEMPLATE_STANDARD),
    "FPT": ("STANDARD", TEMPLATE_STANDARD),
    "TCB": ("BANK", TEMPLATE_BANK),
    "MBB": ("BANK", TEMPLATE_BANK),
    "TCX": ("SECURITIES", TEMPLATE_SECURITIES),
}

NOTE_BY_KIND = {
    "STANDARD": None,
    "BANK": ("Đây là tổ chức tín dụng: báo cáo tài chính không có khái niệm hàng tồn "
             "kho / giá vốn hàng bán / phải thu-phải trả thương mại, và không tách "
             "vay ngắn hạn - dài hạn trên bảng cân đối. Các chỉ tiêu tương ứng hiển "
             "thị \"n/a\". Doanh thu dùng \"Tổng thu nhập hoạt động\" thay cho "
             "\"Doanh thu thuần\"; ROA/Biên lợi nhuận ròng dùng lợi nhuận sau thuế "
             "toàn ngân hàng."),
    "SECURITIES": ("Đây là công ty chứng khoán: \"Giá vốn hàng bán\" được thay bằng "
                   "\"Chi phí hoạt động\", \"Phải thu khách hàng\"/\"Hàng tồn kho\" "
                   "không có số liệu (khoản mục thời kỳ trước 2016, luôn = 0) nên "
                   "Số ngày phải thu và Số ngày tồn kho hiển thị \"n/a\"."),
}

# ---------------------------------------------------------------------------
# 3) Khai báo các chỉ số tài chính (khối CHỈ SỐ TÀI CHÍNH) dưới dạng danh
#    sách khai báo (không phải code mệnh lệnh) để:
#      - đếm được số dòng khối này TRƯỚC khi ghi (dùng để tính trước vị trí
#        dòng của khối DỮ LIỆU TRÍCH XUẤT nằm phía dưới)
#      - áp tô màu theo giá trị (cf) cho từng dòng
# ---------------------------------------------------------------------------

def _m(label, formula_fn, pct=True, indent=False, days=False, cf=None):
    return dict(label=label, formula_fn=formula_fn, pct=pct, indent=indent, days=days, cf=cf)


RATIO_SECTIONS = [
    {
        "title": "1. Khả năng sinh lời",
        "metrics": [
            _m("ROE (LNST CĐ Cty mẹ / VCSH bình quân)",
               lambda R, col, prior: (f"{R('ni_parent', col)}/(({R('equity', col)}+{R('equity', prior)})/2)"
                                       if prior else None), cf="roe"),
            _m("ROA (LNST / Tổng tài sản bình quân)",
               lambda R, col, prior: (f"{R('ni_total', col)}/(({R('total_assets', col)}+{R('total_assets', prior)})/2)"
                                       if prior else None), cf="posneg"),
            _m("Tăng trưởng LN Cổ đông Cty mẹ",
               lambda R, col, prior: (f"({R('ni_parent', col)}-{R('ni_parent', prior)})/ABS({R('ni_parent', prior)})"
                                       if prior else None), cf="growth"),
            _m("Tăng trưởng Doanh thu thuần",
               lambda R, col, prior: (f"({R('revenue', col)}-{R('revenue', prior)})/ABS({R('revenue', prior)})"
                                       if prior else None), cf="growth"),
        ],
    },
    {
        "title": "2. Biên lợi nhuận",
        "metrics": [
            _m("Biên lợi nhuận gộp",
               lambda R, col, prior: f"{R('gross_profit', col)}/{R('revenue', col)}", cf="gross_margin"),
            _m("Biên lợi nhuận ròng",
               lambda R, col, prior: f"{R('ni_total', col)}/{R('revenue', col)}", cf="net_margin"),
        ],
    },
    {
        "title": "3. Đòn bẩy tài chính",
        "metrics": [
            _m("D/E (Nợ vay / Vốn chủ sở hữu)",
               lambda R, col, prior: f"({R('st_borrow', col)}+{R('lt_borrow', col)})/{R('equity', col)}",
               pct=False, cf="debt_equity"),
            _m("Nợ vay dài hạn/Vốn chủ sở hữu",
               lambda R, col, prior: f"{R('lt_borrow', col)}/{R('equity', col)}", pct=False, indent=True),
            _m("Nợ vay ngắn hạn/Vốn chủ sở hữu",
               lambda R, col, prior: f"{R('st_borrow', col)}/{R('equity', col)}", pct=False, indent=True),
            _m("Nợ phải trả/Vốn chủ sở hữu",
               lambda R, col, prior: f"{R('total_liab', col)}/{R('equity', col)}", pct=False, cf="debt_equity"),
            _m("Nợ dài hạn/Vốn chủ sở hữu",
               lambda R, col, prior: f"{R('lt_liab', col)}/{R('equity', col)}", pct=False, indent=True),
            _m("Nợ ngắn hạn/Vốn chủ sở hữu",
               lambda R, col, prior: f"{R('st_liab', col)}/{R('equity', col)}", pct=False, indent=True),
        ],
    },
    {
        "title": "4. Chi phí & khả năng trả lãi",
        "metrics": [
            _m("Chi phí bán hàng, quản lý/LN gộp",
               lambda R, col, prior: f"(ABS({R('selling_exp', col)})+ABS({R('admin_exp', col)}))/{R('gross_profit', col)}",
               cf="sga_ratio"),
            _m("Chi phí lãi vay/LN từ hoạt động kinh doanh",
               lambda R, col, prior: f"ABS({R('interest_exp', col)})/{R('operating_profit', col)}", cf="interest"),
        ],
    },
    {
        "title": "5. Chu kỳ hoạt động vốn lưu động",
        "metrics": [
            _m("Số ngày phải trả (DPO)",
               lambda R, col, prior: (f"(({R('ap', col)}+{R('ap', prior)})/2)/ABS({R('cogs', col)})*365"
                                       if prior else None), days=True, pct=False),
            _m("Số ngày phải thu (DSO)",
               lambda R, col, prior: (f"(({R('ar', col)}+{R('ar', prior)})/2)/{R('revenue', col)}*365"
                                       if prior else None), days=True, pct=False),
            _m("Số ngày tồn kho (DIO)",
               lambda R, col, prior: (f"(({R('inventory', col)}+{R('inventory', prior)})/2)/ABS({R('cogs', col)})*365"
                                       if prior else None), days=True, pct=False),
        ],
    },
    {
        "title": "6. Cơ cấu tài sản",
        "metrics": [
            _m("Tiền và tương đương tiền/Tài sản",
               lambda R, col, prior: f"{R('cash', col)}/{R('total_assets', col)}", cf="cash_ratio"),
            _m("Hàng tồn kho/Tài sản",
               lambda R, col, prior: f"{R('inventory', col)}/{R('total_assets', col)}"),
        ],
    },
]

# Số dòng khối CHỈ SỐ TÀI CHÍNH: mỗi section = 1 dòng tiêu đề + N dòng chỉ
# số. Không còn dòng trắng ngăn cách giữa các nhóm chỉ số (đã bỏ theo yêu
# cầu để bảng gọn hơn).
RATIO_ROWS_COUNT = sum(1 + len(sec["metrics"]) for sec in RATIO_SECTIONS)

# Ngưỡng tô màu theo giá trị (có thể chỉnh lại tuỳ khẩu vị rủi ro).
# Mỗi rule: (operator, formula_list, fill, stopIfTrue).
# - formula_list có 1 phần tử cho các operator 1 vế (lessThan, greaterThan,
#   lessThanOrEqual, greaterThanOrEqual...), có 2 phần tử [min, max] cho
#   operator 2 vế (between/notBetween).
# - stopIfTrue=True dùng cho rule "đỏ" ở ranh giới để tránh chồng lấn màu
#   với rule "vàng"/"xanh" liền kề khi giá trị đúng bằng mốc biên.
CF_THRESHOLDS = {
    # ROE: < 0.15 đỏ, >= 0.15 xanh lá
    "roe": [
        ("lessThan", ["0.15"], FILL_BAD),
        ("greaterThanOrEqual", ["0.15"], FILL_GOOD),
    ],
    # Tăng trưởng LN sau thuế / Tăng trưởng DT: > 0.15 xanh lá, < 0.1 đỏ
    "growth": [
        ("greaterThan", ["0.15"], FILL_GOOD),
        ("lessThan", ["0.1"], FILL_BAD),
    ],
    # Biên lợi nhuận gộp: <=0.1 đỏ, 0.1-0.3 vàng, >=0.3 xanh lá
    "gross_margin": [
        ("lessThanOrEqual", ["0.1"], FILL_BAD, True),
        ("between", ["0.1", "0.3"], FILL_AMBER),
        ("greaterThanOrEqual", ["0.3"], FILL_GOOD),
    ],
    # Biên lợi nhuận ròng: >= 0.1 xanh lá (không có ngưỡng đỏ theo yêu cầu)
    "net_margin": [
        ("greaterThanOrEqual", ["0.1"], FILL_GOOD),
    ],
    # D/E, Nợ phải trả/VCSH: <= 0.5 xanh lá (không có ngưỡng đỏ theo yêu cầu)
    "debt_equity": [
        ("lessThanOrEqual", ["0.5"], FILL_GOOD),
    ],
    # Chi phí bán hàng, quản lý/LN gộp: > 0.7 đỏ, <= 0.3 xanh lá
    "sga_ratio": [
        ("greaterThan", ["0.7"], FILL_BAD),
        ("lessThanOrEqual", ["0.3"], FILL_GOOD),
    ],
    # Tiền mặt và tương đương/TS: ngoài [0.02, 0.3] đỏ, trong khoảng xanh lá
    "cash_ratio": [
        ("notBetween", ["0.02", "0.3"], FILL_BAD),
        ("between", ["0.02", "0.3"], FILL_GOOD),
    ],
    # Các rule cũ, giữ lại cho những chỉ tiêu không nằm trong bảng yêu cầu mới
    # (ROA, Nợ vay dài/ngắn hạn từng phần, Chi phí lãi vay/LN HĐKD...).
    "posneg": [("greaterThan", ["0"], FILL_GOOD), ("lessThan", ["0"], FILL_BAD)],
    "leverage": [("greaterThan", ["1"], FILL_BAD), ("lessThan", ["0.3"], FILL_GOOD)],
    "interest": [("greaterThan", ["0.5"], FILL_BAD), ("lessThan", ["0.15"], FILL_GOOD)],
}


def _apply_cf(ws, row, cf_key, last_col):
    rng = f"B{row}:{last_col}{row}"
    for rule in CF_THRESHOLDS[cf_key]:
        operator, formula, fill = rule[0], rule[1], rule[2]
        stop_if_true = rule[3] if len(rule) > 3 else False
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator=operator, formula=formula, fill=fill,
                       stopIfTrue=stop_if_true or None),
        )


def build_workbook_ratio_sheet(path, symbol):
    kind, template = SYMBOL_TEMPLATE[symbol]

    wb = openpyxl.load_workbook(path, data_only=False)

    # ---- Xác định số năm THỰC TẾ từ dòng 1 sheet balance_sheet (không cố
    # định 4 cột như trước) -> không bỏ sót năm nào khi có dữ liệu mới. ----
    bs_ws = wb["balance_sheet"]
    n_years = 0
    while bs_ws.cell(1, 2 + n_years).value not in (None, ""):
        n_years += 1
    if n_years == 0:
        n_years = 4
    YEAR_COLS = [get_column_letter(2 + i) for i in range(n_years)]
    # Xếp TĂNG DẦN trái->phải (cũ nhất bên trái, mới nhất bên phải), khớp
    # đúng thứ tự cột nguồn -> không cần đảo cột khi lấy header năm.
    REVERSE_COL = {col: col for col in YEAR_COLS}
    PRIOR_COL = {col: (YEAR_COLS[i - 1] if i > 0 else None) for i, col in enumerate(YEAR_COLS)}
    last_col = YEAR_COLS[-1]
    last_col_idx = column_index_from_string(last_col)
    spacer_col = get_column_letter(last_col_idx + 1)
    n_cols = last_col_idx  # số cột A..last_col dùng để merge/fill header

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    # chèn làm sheet thứ 4 (sau balance_sheet, income_statement, cash_flow)
    idx = min(3, len(wb.sheetnames))
    ws = wb.create_sheet(SHEET_NAME, idx)

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions[spacer_col].width = 3

    r = 1
    c = ws.cell(r, 1, f"CHỈ SỐ TÀI CHÍNH — {symbol}")
    c.font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    for col in range(1, n_cols + 1):
        ws.cell(r, col).fill = HEADER_FILL
    ws.row_dimensions[r].height = 22

    # (Không còn chú thích phương pháp luận / lưu ý loại hình DN, không còn
    # dòng trắng ở đầu sheet theo yêu cầu — đi thẳng vào bảng chỉ số.)
    r += 1

    # -------------------------------------------------------------------
    # Tính trước vị trí dòng của khối II (DỮ LIỆU TRÍCH XUẤT, nằm dưới)
    # để khối I (CHỈ SỐ TÀI CHÍNH, nằm trên) có thể tham chiếu công thức
    # tới đúng số dòng ngay từ đầu.
    # -------------------------------------------------------------------
    ratio_section_row = r
    ratio_header_row = ratio_section_row + 1
    raw_section_row = ratio_header_row + 1 + RATIO_ROWS_COUNT + 1  # +1 dòng trắng ngăn 2 khối
    raw_header_row = raw_section_row + 1
    RAW_ROW = {key: raw_header_row + 1 + i for i, (key, _) in enumerate(RAW_ITEMS)}

    def R(key, col):
        """Ô dữ liệu thô (khối II) tại cột col, dùng $ cho dòng để copy công thức."""
        return f"{col}${RAW_ROW[key]}"

    # ---------------- KHỐI I: CHỈ SỐ TÀI CHÍNH (đưa lên trên) ----------------
    c = ws.cell(ratio_section_row, 1, "I. CHỈ SỐ TÀI CHÍNH")
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    for col in range(1, n_cols + 1):
        ws.cell(ratio_section_row, col).fill = SECTION_FILL

    ws.cell(ratio_header_row, 1, "Chỉ tiêu").font = Font(name=FONT_NAME, bold=True)
    for col in YEAR_COLS:
        src_col = REVERSE_COL[col]
        cell = ws.cell(ratio_header_row, column_index_from_string(col), f"='balance_sheet'!{src_col}1")
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, n_cols + 1):
        ws.cell(ratio_header_row, col).border = BORDER

    r = ratio_header_row + 1

    def section(title):
        nonlocal r
        c = ws.cell(r, 1, title)
        c.font = Font(name=FONT_NAME, size=10, bold=True, italic=True)
        r += 1

    def metric(label, formula_fn, pct=True, indent=False, days=False, cf=None):
        nonlocal r
        lbl_cell = ws.cell(r, 1, ("     - " if indent else "") + label)
        lbl_cell.font = Font(name=FONT_NAME, size=10, bold=not indent)
        lbl_cell.border = BORDER
        for col in YEAR_COLS:
            prior = PRIOR_COL[col]
            cell = ws.cell(r, column_index_from_string(col))
            expr = formula_fn(R, col, prior)
            if expr is None:
                cell.value = "n/a"
                cell.font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)
            else:
                cell.value = f'=IFERROR({expr},"n/a")'
                cell.font = Font(name=FONT_NAME, size=10, color=BLACK)
                if days:
                    cell.number_format = "#,##0 \"ngày\""
                elif pct:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "0.00\"x\""
            cell.border = BORDER
        if cf:
            _apply_cf(ws, r, cf, last_col)
        r += 1

    # Không còn dòng trắng ngăn cách giữa các nhóm chỉ số (đã bỏ theo yêu cầu).
    for sec in RATIO_SECTIONS:
        section(sec["title"])
        for m in sec["metrics"]:
            metric(**m)

    r += 1  # 1 dòng trắng ngăn khối I và khối II

    assert r == raw_section_row, (r, raw_section_row)  # kiểm tra layout tính trước khớp thực tế

    # ---------------- KHỐI II: DỮ LIỆU TRÍCH XUẤT (đẩy xuống dưới) ----------
    c = ws.cell(raw_section_row, 1, "II. DỮ LIỆU TRÍCH XUẤT (tự động dò theo tên khoản mục)")
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    for col in range(1, n_cols + 1):
        ws.cell(raw_section_row, col).fill = SECTION_FILL

    ws.cell(raw_header_row, 1, "Khoản mục").font = Font(name=FONT_NAME, bold=True)
    for col in YEAR_COLS:
        src_col = REVERSE_COL[col]
        cell = ws.cell(raw_header_row, column_index_from_string(col), f"='balance_sheet'!{src_col}1")
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, n_cols + 1):
        ws.cell(raw_header_row, col).border = BORDER

    for key, default_label in RAW_ITEMS:
        row = RAW_ROW[key]
        mapping = template.get(key)
        # Nhãn hiển thị lấy ĐÚNG theo tên khoản mục gốc trong báo cáo của
        # công ty (mapping[1]); chỉ dùng nhãn chuẩn hoá mặc định khi khoản
        # mục này không tồn tại với loại hình DN đó (mapping is None).
        display_label = mapping[1] if mapping else default_label
        lbl_cell = ws.cell(row, 1, display_label)
        lbl_cell.font = Font(name=FONT_NAME, size=10)
        lbl_cell.border = BORDER
        for col in YEAR_COLS:
            cell = ws.cell(row, column_index_from_string(col))
            if mapping is None:
                cell.value = "n/a"
                cell.font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)
                cell.comment = Comment(
                    "Không có khoản mục tương ứng trong báo cáo tài chính của công ty này.",
                    "auto")
            else:
                sheet_name, item_label = mapping
                item_label_esc = item_label.replace('"', '""')
                formula = (
                    f'=IFERROR(INDEX({sheet_name}!$B:${last_col},'
                    f'MATCH("{item_label_esc}",{sheet_name}!$A:$A,0),'
                    f'MATCH({col}${raw_header_row},{sheet_name}!$B$1:${last_col}$1,0)),"")'
                )
                cell.value = formula
                cell.font = Font(name=FONT_NAME, size=10, color=GREEN)
                cell.number_format = "#,##0;(#,##0);\"-\""
            cell.border = BORDER

    ws.freeze_panes = "B" + str(ratio_header_row + 1)

    wb.save(path)
    return path


def main():
    for symbol in SYMBOL_TEMPLATE:
        path = os.path.join(BASE_DIR, symbol, f"{symbol}_financials.xlsx")
        if not os.path.exists(path):
            print("MISSING", path)
            continue
        build_workbook_ratio_sheet(path, symbol)
        print("OK", symbol)


if __name__ == "__main__":
    main()
