"""
Gộp 3 file CSV (balance_sheet, income_statement, cash_flow) của mỗi mã trong
thư mục financials/<MÃ>/ thành 1 file Excel financials/<MÃ>/<MÃ>_financials.xlsx
với 3 sheet cùng tên.

Cách gộp dữ liệu qua thời gian (để không mất các năm cũ khi vnstock chỉ trả
về vài năm gần nhất mỗi lần fetch):
    - Dữ liệu mới được đối chiếu với dữ liệu cũ đã có trong file Excel theo
      cột "item" (tên khoản mục).
    - Nếu "item" TRÙNG KHỚP với 1 dòng đã có -> merge các cột năm mới vào
      dòng đó (đè giá trị nếu năm đã tồn tại, thêm cột nếu là năm mới) ->
      qua nhiều lần chạy sẽ tích luỹ dữ liệu nhiều năm.
    - Nếu "item" KHÔNG TRÙNG KHỚP (khoản mục hoàn toàn mới, chưa từng có
      trong dữ liệu cũ) -> KHÔNG merge dòng đó, dữ liệu cũ giữ nguyên.
    - Nếu file Excel chưa tồn tại (lần chạy đầu tiên) -> dùng toàn bộ CSV
      hiện tại làm dữ liệu nền ban đầu.

Nếu file Excel đã có thêm sheet khác (vd "financial_ratios" tự thêm sau
này), sheet đó được giữ nguyên, không bị đụng tới.

Chạy:
    python merge_financials.py                    # gộp tất cả mã có trong financials/
    python merge_financials.py HPG                 # 1 mã
    python merge_financials.py HPG,TCB,FPT,PNJ      # nhiều mã, cách nhau dấu phẩy
"""

import sys
import os
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FINANCIALS_DIR = "financials"

FONT_NAME = "Tahoma"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tên sheet <-> hậu tố tên file CSV, theo đúng thứ tự hiển thị trong Excel.
STATEMENT_SHEETS = {
    "balance_sheet": "balance_sheet",
    "income_statement": "income_statement",
    "cash_flow": "cash_flow",
}

DROP_COLUMNS = ["item_en", "item_id"]


def _load_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, encoding="utf-8-sig")
    drop_cols = [c for c in DROP_COLUMNS if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    return df


def _sort_year_columns(cols):
    """Sắp xếp các cột năm giảm dần (mới nhất trước), cột không phải số năm
    (nếu có) được đẩy xuống cuối, giữ nguyên thứ tự gốc."""

    def key(c):
        try:
            return (0, -int(c))
        except (TypeError, ValueError):
            return (1, str(c))

    return sorted(cols, key=key)


def merge_sheet(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """Merge new_df vào old_df theo cột 'item'. Xem quy tắc merge ở đầu file."""
    if old_df is None or old_df.empty:
        # Chưa có dữ liệu cũ -> dùng dữ liệu mới làm nền.
        merged = new_df.copy()
    else:
        merged = old_df.copy()
        year_cols_new = [c for c in new_df.columns if c != "item"]

        # Đảm bảo các cột năm mới tồn tại trong merged (dù chưa có dòng nào khớp).
        for yc in year_cols_new:
            if yc not in merged.columns:
                merged[yc] = pd.NA

        item_to_index = {}
        for idx, item_val in merged["item"].items():
            item_to_index.setdefault(item_val, idx)

        for _, row in new_df.iterrows():
            item_val = row["item"]
            if item_val not in item_to_index:
                # Không trùng khớp -> bỏ qua, không merge, giữ nguyên dữ liệu cũ.
                continue
            old_idx = item_to_index[item_val]
            for yc in year_cols_new:
                merged.loc[old_idx, yc] = row[yc]

    year_cols = _sort_year_columns([c for c in merged.columns if c != "item"])
    merged = merged[["item"] + year_cols]
    return merged


def _format_statement_sheet(ws) -> None:
    """Định dạng chuẩn cho 1 sheet báo cáo (balance_sheet/income_statement/
    cash_flow) vừa được pandas ghi ra: font Tahoma, dòng tiêu đề (năm) in
    đậm/nền màu, số có dấu phân cách nghìn, cố định dòng 1 + cột A, độ rộng
    cột hợp lý, viền mảnh cho toàn vùng dữ liệu."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = BORDER
            if cell.row == 1:
                cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left")
            else:
                cell.font = Font(name=FONT_NAME, size=10)
                if cell.column > 1:
                    cell.number_format = "#,##0;(#,##0);\"-\""

    ws.freeze_panes = "B2"


def process_symbol(symbol: str) -> bool:
    sym_dir = os.path.join(FINANCIALS_DIR, symbol)
    out_path = os.path.join(sym_dir, f"{symbol}_financials.xlsx")

    if not os.path.isdir(sym_dir):
        print(f"  Bỏ qua {symbol}: không tìm thấy thư mục {sym_dir}")
        return False

    # Chỉ đọc bằng pandas 3 sheet báo cáo gốc (dữ liệu thuần, không công thức).
    # KHÔNG dùng pd.read_excel cho toàn bộ workbook: pandas chỉ đọc được giá
    # trị đã tính sẵn (cached value) của ô công thức chứ không đọc được công
    # thức, nên nếu nạp rồi ghi lại các sheet khác (vd "chi_so_tai_chinh")
    # qua pandas thì mọi công thức trong đó sẽ bị "phẳng hoá" thành số tĩnh.
    # Các sheet ngoài 3 sheet báo cáo được giữ nguyên 100% (kể cả công thức,
    # định dạng) bằng cách copy trực tiếp qua openpyxl ở bước bên dưới.
    existing_sheets = {}
    other_sheet_names = []
    if os.path.exists(out_path):
        try:
            existing_sheets = pd.read_excel(
                out_path, sheet_name=list(STATEMENT_SHEETS.keys()), engine="openpyxl")
        except Exception as e:
            print(f"  CẢNH BÁO: không đọc được file Excel cũ {out_path} ({e}) -> tạo mới.")
            existing_sheets = {}
        try:
            other_sheet_names = [
                s for s in openpyxl.load_workbook(out_path, read_only=True).sheetnames
                if s not in STATEMENT_SHEETS
            ]
        except Exception:
            other_sheet_names = []

    backup_path = out_path + ".bak_other_sheets.xlsx"
    if other_sheet_names and os.path.exists(out_path):
        shutil.copyfile(out_path, backup_path)

    output_sheets = {}
    any_written = False

    for sheet_name, suffix in STATEMENT_SHEETS.items():
        csv_path = os.path.join(sym_dir, f"{symbol}_{suffix}.csv")
        if not os.path.exists(csv_path):
            print(f"  Bỏ qua sheet '{sheet_name}': không tìm thấy {csv_path}")
            continue

        new_df = _load_csv(csv_path)
        old_df = existing_sheets.get(sheet_name)
        output_sheets[sheet_name] = merge_sheet(old_df, new_df)
        any_written = True

    if not any_written:
        print(f"  Không có CSV nào cho {symbol}, bỏ qua.")
        return False

    ordered_names = [n for n in STATEMENT_SHEETS if n in output_sheets]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name in ordered_names:
            output_sheets[name].to_excel(writer, sheet_name=name, index=False)
            _format_statement_sheet(writer.sheets[name])

    # Copy nguyên trạng (công thức + định dạng) các sheet khác từ file cũ
    # (vd "chi_so_tai_chinh") sang file vừa ghi ở trên.
    final_order = list(ordered_names)
    if other_sheet_names:
        _copy_other_sheets(out_path, other_sheet_names)
        final_order += other_sheet_names

    print(f"  Đã ghi {out_path}  (sheets: {', '.join(final_order)})")
    return True


def _copy_other_sheets(out_path, sheet_names):
    """Copy nguyên trạng (giá trị, công thức, style cơ bản, độ rộng cột) các
    sheet có tên trong `sheet_names` từ BẢN CŨ (backup tạm lấy trước khi
    pandas ghi đè) sang file `out_path` vừa được pandas ghi lại (chỉ chứa 3
    sheet báo cáo)."""
    backup_path = out_path + ".bak_other_sheets.xlsx"
    if not os.path.exists(backup_path):
        return
    src_wb = openpyxl.load_workbook(backup_path, data_only=False)
    dst_wb = openpyxl.load_workbook(out_path)
    for name in sheet_names:
        if name not in src_wb.sheetnames:
            continue
        src_ws = src_wb[name]
        dst_ws = dst_wb.create_sheet(name)
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
        for col, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col].width = dim.width
        for row_dim_idx, row_dim in src_ws.row_dimensions.items():
            if row_dim.height:
                dst_ws.row_dimensions[row_dim_idx].height = row_dim.height
        for merged_range in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged_range))
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format
        if src_ws.freeze_panes:
            dst_ws.freeze_panes = src_ws.freeze_panes
    dst_wb.save(out_path)
    os.remove(backup_path)


def main():
    if len(sys.argv) > 1:
        symbols = [s.strip().upper() for s in sys.argv[1].split(",") if s.strip()]
    else:
        # Mặc định: gộp tất cả các mã đang có sẵn trong financials/
        if os.path.isdir(FINANCIALS_DIR):
            symbols = sorted(
                d for d in os.listdir(FINANCIALS_DIR)
                if os.path.isdir(os.path.join(FINANCIALS_DIR, d))
            )
        else:
            symbols = []

    if not symbols:
        print(f"Không tìm thấy mã nào trong '{FINANCIALS_DIR}/'.")
        sys.exit(1)

    any_success = False
    for symbol in symbols:
        print(f"\n=== Mã {symbol} ===")
        if process_symbol(symbol):
            any_success = True

    if not any_success:
        print("\nKhông gộp được dữ liệu cho bất kỳ mã nào.")
        sys.exit(1)

    print("\nHoàn tất.")


if __name__ == "__main__":
    main()
